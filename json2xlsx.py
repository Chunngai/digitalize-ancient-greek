import argparse
import json
import time

import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Side, Border

FONT = Font(name="Calibri")
CENTER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center"
)
SIDE = Side(border_style="thin", color="000000")
BORDER = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)

LIGHT_BLUE_FILL = PatternFill(patternType="solid", fgColor="DDEBF7")
DARK_BLUE_FILL = PatternFill(patternType="solid", fgColor="BDD7EE")
LIGHT_ORANGE_FILL = PatternFill(patternType="solid", fgColor="FDE9D9")
DARK_ORANGE_FILL = PatternFill(patternType="solid", fgColor="FCD5B4")

ORANGE_FILLS = [
    LIGHT_ORANGE_FILL,
    DARK_ORANGE_FILL,
]


def read_json(fp_json):
    with open(fp_json, "r", encoding="utf-8") as f:
        return json.load(f)


def set_vertical_alignment(sheet, cols="ABCDEF", rows=range(1, 200)):
    for col in cols:
        for row in rows:
            sheet[f"{col}{row}"].alignment = Alignment(vertical="center")


def set_column_widths(sheet, cols="ABCDEF", width=15):
    for column_id in cols:
        sheet.column_dimensions[column_id].width = width


def set_font(sheet, cols="ABCDEF", rows=range(1, 200)):
    for col in cols:
        for row in rows:
            sheet[f"{col}{row}"].font = FONT


def set_format(sheet, cols="ABCDEF", rows=range(1, 200)):
    for col in cols:
        for row in rows:
            sheet[f"{col}{row}"].number_format = "@"


def set_border(sheet, cols="ABCDEF", rows=range(1, 200)):
    for col in cols:
        for row in rows:
            sheet[f"{col}{row}"].border = BORDER


def make_title(sheet, loc, text, fill=LIGHT_BLUE_FILL):
    sheet[loc] = text
    sheet[loc].font = FONT
    sheet[loc].alignment = CENTER_ALIGNMENT
    sheet[loc].fill = fill


def make_index(sheet, loc, index, fill=None):
    sheet[loc] = str(index)
    sheet[loc].font = FONT
    sheet[loc].alignment = CENTER_ALIGNMENT
    if fill is not None:
        sheet[loc].fill = fill


def get_item(d, k, default=""):
    try:
        return d[k]
    except KeyError:
        return default


def get_list(d, k, n):
    try:
        return d[k]
    except KeyError:
        return [
            {}
            for _ in range(n)
        ]


def json2xlsx(lesson_json):
    def toggle_orange_fill_index():
        nonlocal orange_fill_idx
        orange_fill_idx = 1 - orange_fill_idx

    def next_row():
        nonlocal row_idx
        row_idx += 1

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    set_vertical_alignment(sheet)
    set_column_widths(sheet)
    set_font(sheet)
    set_format(sheet)

    row_idx = 0
    orange_fill_idx = 0

    # The first row.
    next_row()

    # ID label.
    make_title(sheet, f"A{row_idx}", "ID")
    # ID content.
    sheet[f"B{row_idx}"] = get_item(lesson_json, "id")
    sheet[f"B{row_idx}"].alignment = CENTER_ALIGNMENT

    # Fields.
    make_title(sheet, f"C{row_idx}", "STM/POS/WRD")
    make_title(sheet, f"D{row_idx}", "PRS/MNS/EXP")
    make_title(sheet, f"E{row_idx}", "SFS/LBS/EXP")
    make_title(sheet, f"F{row_idx}", "ARS/USG/EXP")

    # Title.
    next_row()
    # Title label.
    make_title(sheet, f"A{row_idx}", "TITLE")
    # Title content.
    sheet[f"B{row_idx}"] = get_item(lesson_json, "title")
    sheet[f"B{row_idx}"].alignment = CENTER_ALIGNMENT
    sheet.merge_cells(f"B{row_idx}:F{row_idx}")

    # Vocab.
    vocab_content_start_row = row_idx + 1
    vocab = get_list(lesson_json, "vocab", 10)
    for word in vocab:
        next_row()

        make_title(sheet, f"B{row_idx}", "FORMS", fill=ORANGE_FILLS[orange_fill_idx])
        stem_list = []
        prefixes_list = []
        suffices_list = []
        articles_list = []
        for form in get_list(word, "forms", 1):
            stem_list.append(get_item(form, "stem", "-"))
            prefixes_list.append(",".join(get_item(form, "prefixes", "-")))
            suffices_list.append(",".join(get_item(form, "suffices", "-")))
            articles_list.append(",".join(get_item(form, "articles", "-")))
        sheet[f"C{row_idx}"] = "\n".join(stem_list)
        sheet[f"D{row_idx}"] = "\n".join(prefixes_list)
        sheet[f"E{row_idx}"] = "\n".join(suffices_list)
        sheet[f"F{row_idx}"] = "\n".join(articles_list)

        next_row()

        make_title(sheet, f"B{row_idx}", "MEANINGS", fill=ORANGE_FILLS[orange_fill_idx])
        pos_list = []
        meanings_list = []
        labels_list = []
        usage_list = []
        for meanings in get_list(word, "meanings", 1):
            pos_list.append(get_item(meanings, "pos", "-"))
            meanings_list.append(get_item(meanings, "meanings", "-"))
            labels_list.append(",".join(get_item(meanings, "labels", "-")))
            usage_list.append(",".join(get_item(meanings, "usage", "-")))
        sheet[f"C{row_idx}"] = "\n".join(pos_list)
        sheet[f"D{row_idx}"] = "\n".join(meanings_list)
        sheet[f"E{row_idx}"] = "\n".join(labels_list)
        sheet[f"F{row_idx}"] = "\n".join(usage_list)

        next_row()

        make_title(sheet, f"B{row_idx}", "EXPLANATION", fill=ORANGE_FILLS[orange_fill_idx])
        sheet[f"C{row_idx}"] = get_item(word, "explanation")
        sheet.merge_cells(f"C{row_idx}:F{row_idx}")

        toggle_orange_fill_index()
    # Vocab title.
    make_title(sheet, f"A{vocab_content_start_row}", "VOCABULARY")
    sheet.merge_cells(f"A{vocab_content_start_row}:A{vocab_content_start_row + row_idx - vocab_content_start_row}")

    # Sentences.
    sentences_content_start_row = row_idx + 1
    sentences = get_list(lesson_json, "sentences", 15)
    for sentence_idx, sentence in enumerate(sentences):
        next_row()

        make_title(sheet, f"B{row_idx}", f"{'GREEK' if sentence_idx < 10 else 'ENGLISH'}",
                   fill=ORANGE_FILLS[orange_fill_idx])
        sheet[f"C{row_idx}"] = (
            sentence["greek" if "greek" in sentence.keys() else "english"]
            if sentence != {}
            else ""
        )
        sheet.merge_cells(f"C{row_idx}:F{row_idx}")

        next_row()

        make_title(sheet, f"B{row_idx}", f"{'ENGLISH_' if sentence_idx < 10 else 'GREEK_'}",
                   fill=ORANGE_FILLS[orange_fill_idx])
        sheet[f"C{row_idx}"] = (
            sentence["english_" if "english_" in sentence.keys() else "greek_"]
            if sentence != {}
            else ""
        )
        sheet.merge_cells(f"C{row_idx}:F{row_idx}")

        toggle_orange_fill_index()
    # Sentences title.
    make_title(sheet, f"A{sentences_content_start_row}", "SENTENCES")
    sheet.merge_cells(
        f"A{sentences_content_start_row}:A{sentences_content_start_row + row_idx - sentences_content_start_row}")

    # Reading.
    reading_start_row = row_idx + 1
    reading = get_item(lesson_json, "reading", default={})

    # Title.
    next_row()
    # Title label.
    make_title(sheet, f"B{row_idx}", "TITLE")
    # Title content.
    sheet[f"C{row_idx}"] = get_item(reading, "title")
    sheet.merge_cells(f"C{row_idx}:F{row_idx}")

    # Text.
    next_row()
    # Text label.
    make_title(sheet, f"B{row_idx}", "TEXT")
    # Text content.
    sheet[f"C{row_idx}"] = get_item(reading, "text")
    sheet.merge_cells(f"C{row_idx}:F{row_idx}")

    # Vocab content.
    reading_vocab_content_start_row = row_idx + 1
    reading_vocab = get_list(reading, "vocab", 10)
    for word_idx, word in enumerate(reading_vocab):
        next_row()

        sheet[f"C{row_idx}"] = get_item(word, "word")
        sheet[f"D{row_idx}"] = get_item(word, "explanation")
        sheet.merge_cells(f"D{row_idx}:F{row_idx}")

        toggle_orange_fill_index()
    # Vocab label.
    make_title(sheet, f"B{reading_vocab_content_start_row}", "VOCABULARY")
    sheet.merge_cells(
        f"B{reading_vocab_content_start_row}:B{reading_vocab_content_start_row + row_idx - reading_vocab_content_start_row}")

    # Translation.
    next_row()
    # Translation label.
    make_title(sheet, f"B{row_idx}", "TRANSLATION")
    # Translation content.
    sheet[f"C{row_idx}"] = get_item(reading, "translation")
    sheet.merge_cells(f"C{row_idx}:F{row_idx}")

    # Reading label.
    make_title(sheet, f"A{reading_start_row}", "READING")
    sheet.merge_cells(f"A{reading_start_row}:A{reading_start_row + row_idx - reading_start_row}")

    set_border(sheet, rows=range(1, row_idx + 1))

    some_long_vowels = ["ᾱ́", "ᾱ̀", "ῑ́", "ῑ̀", "ῡ́", "ῡ̀"]
    for vowel, col in zip(some_long_vowels, "HIJKLM"):
        sheet[f"{col}1"] = vowel

    return workbook


def main(fp_json, fp_xlsx):
    try:
        lesson_json = read_json(fp_json=fp_json)
    except FileNotFoundError:
        lesson_json = {}
    lesson_xlsx = json2xlsx(lesson_json=lesson_json)
    lesson_xlsx.save(fp_xlsx)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("--json", nargs=1)
    parser.add_argument("--new", action="store_true")
    parser.add_argument("--xlsx")
    args = parser.parse_args()

    if args.json:
        assert args.new is None
    if args.new:
        assert args.json is None
        args.json = ""

    if not args.xlsx:
        args.xlsx = args.json.replace(".json", f"_{time.strftime('%y%m%d%H%M%S')}.xlsx")

    main(fp_json=args.json, fp_xlsx=args.xlsx)
